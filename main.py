import io
import os
import base64
import re
import json
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Optional
from contextlib import asynccontextmanager

from fastapi import FastAPI, Request, Form, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from fastapi.middleware.cors import CORSMiddleware
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired

import database
import parser as xlsx_parser

# ── Gmail auto-import ─────────────────────────────────────────────────────────

GMAIL_SENDER   = "rmk@organicmarket.dk"
GMAIL_SCOPES   = ["https://www.googleapis.com/auth/gmail.readonly"]


def _gmail_creds():
    """Byg Gmail Credentials fra GMAIL_TOKEN_JSON env var."""
    token_json = os.environ.get("GMAIL_TOKEN_JSON", "")
    if not token_json:
        raise RuntimeError("GMAIL_TOKEN_JSON env var ikke sat")
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request as GRequest
    info = json.loads(token_json)
    creds = Credentials.from_authorized_user_info(info, GMAIL_SCOPES)
    if not creds.valid and creds.expired and creds.refresh_token:
        creds.refresh(GRequest())
    return creds


def _iter_parts(payload):
    yield payload
    for part in payload.get("parts", []):
        yield from _iter_parts(part)


def _tal(s) -> float:
    if s is None: return 0.0
    s = re.sub(r"[^\d,.\-]", "", str(s).strip())
    if not s or s == "-": return 0.0
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        parts = s.lstrip("-").split(".")
        if not (len(parts) == 2 and len(parts[1]) in (1, 2)):
            s = s.replace(".", "")
    try: return float(s)
    except: return 0.0


def _sidst_tal(linje: str) -> float:
    tal = re.findall(r"-?[0-9]{1,3}(?:\.[0-9]{3})*(?:,[0-9]+)?|-?[0-9]+(?:,[0-9]+)?", linje)
    return abs(_tal(tal[-1])) if tal else 0.0


def _parse_faktura_tekst(tekst: str, uge: int, aar: int) -> dict:
    retur_wiener = retur_boller = tgtg = b_kvali = faktura = subtotal = 0.0
    _next_levering = False
    for linje in tekst.splitlines():
        l = linje.strip(); ll = l.lower()
        if _next_levering:
            val = _sidst_tal(l)
            if val > 100: faktura = val
            _next_levering = False
            continue
        if re.search(r"retur\s+wien", ll):
            retur_wiener = _sidst_tal(l)
        elif re.search(r"retur\s+boller", ll):
            retur_boller = _sidst_tal(l)
        elif re.search(r"tgtg|too\s*good\s*to\s*go", ll):
            tgtg = _sidst_tal(l)
        elif re.search(r"kvali|b-?kredit|kvalitets", ll):
            if not re.search(r"retur|wiener|boller|tgtg", ll):
                b_kvali = _sidst_tal(l)
        elif re.search(r"levering", ll):
            val = _sidst_tal(l)
            if val > 100: faktura = val
            else: _next_levering = True
        elif re.search(r"subtotal", ll):
            subtotal = _sidst_tal(l)
        elif not faktura and re.search(r"total\s+dkk\s*:", ll):
            faktura = _sidst_tal(l)
    retur_ialt = round(retur_wiener + retur_boller + tgtg + b_kvali, 2)
    if not faktura and subtotal > 0:
        faktura = round(subtotal + retur_ialt, 2)
    return {
        "uge": uge, "aar": aar,
        "retur_wiener": round(retur_wiener, 2),
        "retur_boller": round(retur_boller, 2),
        "tgtg": round(tgtg, 2),
        "b_kvali": round(b_kvali, 2),
        "retur_ialt": retur_ialt,
        "faktura": round(faktura, 2),
    }


def gmail_sync_run() -> dict:
    """Hent nye bager-fakturaer fra Gmail og gem i databasen. Returnerer status-dict."""
    try:
        from googleapiclient.discovery import build
        import pdfplumber
    except ImportError as e:
        msg = f"Mangler pakke: {e}"
        database.log_gmail_sync("fejl", msg)
        return {"ok": False, "besked": msg}

    try:
        creds = _gmail_creds()
    except Exception as e:
        msg = f"Gmail credentials fejl: {e}"
        database.log_gmail_sync("fejl", msg)
        return {"ok": False, "besked": msg}

    try:
        service  = build("gmail", "v1", credentials=creds)
        allerede = database.hent_gmail_importerede()

        query  = f"from:{GMAIL_SENDER} has:attachment filename:pdf"
        result = service.users().messages().list(userId="me", q=query, maxResults=50).execute()
        messages = result.get("messages", [])

        nye = []
        for msg in messages:
            msg_id = msg["id"]
            if msg_id in allerede:
                continue
            full    = service.users().messages().get(userId="me", id=msg_id, format="full").execute()
            headers = {h["name"]: h["value"] for h in full["payload"]["headers"]}
            subject = headers.get("Subject", "")
            date_str= headers.get("Date", "")

            m_uge = re.search(r"uge\s*(\d+)", subject, re.IGNORECASE)
            if not m_uge:
                continue
            uge = int(m_uge.group(1))
            m_aar = re.search(r"(202\d)", date_str)
            aar = int(m_aar.group(1)) if m_aar else datetime.now().year

            att_id = att_name = None
            for part in _iter_parts(full["payload"]):
                if part.get("mimeType") == "application/pdf":
                    att_id   = part["body"].get("attachmentId")
                    att_name = part.get("filename", "faktura.pdf")
                    break
            if not att_id:
                continue

            att   = service.users().messages().attachments().get(userId="me", messageId=msg_id, id=att_id).execute()
            pdf_b = base64.urlsafe_b64decode(att.get("data", "") + "==")

            tekst = ""
            with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                tmp.write(pdf_b); tmp_path = tmp.name
            try:
                with pdfplumber.open(tmp_path) as pdf:
                    for page in pdf.pages:
                        tekst += (page.extract_text() or "") + "\n"
            finally:
                os.unlink(tmp_path)

            parsed = _parse_faktura_tekst(tekst, uge, aar)
            nye.append({"msg_id": msg_id, "data": parsed})

        if not nye:
            database.log_gmail_sync("ingen_nye", "Ingen nye fakturaer fundet", 0)
            return {"ok": True, "besked": "Ingen nye fakturaer", "antal": 0}

        # Gem i database
        database.gem_bager_regnskab([e["data"] for e in nye])
        for entry in nye:
            database.gem_gmail_importeret(entry["msg_id"], entry["data"]["uge"], entry["data"]["aar"])

        besked = f"Importeret {len(nye)} faktura(er): " + ", ".join(f"uge {e['data']['uge']}/{e['data']['aar']}" for e in nye)
        database.log_gmail_sync("ok", besked, len(nye))
        return {"ok": True, "besked": besked, "antal": len(nye)}

    except Exception as e:
        msg = f"Sync fejl: {e}"
        database.log_gmail_sync("fejl", msg)
        return {"ok": False, "besked": msg}


# ── APScheduler ───────────────────────────────────────────────────────────────

from apscheduler.schedulers.asyncio import AsyncIOScheduler

_scheduler = AsyncIOScheduler(timezone="Europe/Copenhagen")

def _planlagt_gmail_sync():
    import asyncio
    loop = asyncio.new_event_loop()
    try:
        result = gmail_sync_run()
        print(f"[Gmail auto-sync] {result.get('besked','?')}")
    finally:
        loop.close()


@asynccontextmanager
async def lifespan(app: FastAPI):
    database.init_db()
    if os.environ.get("GMAIL_TOKEN_JSON"):
        _scheduler.add_job(_planlagt_gmail_sync, "cron", day_of_week="mon,thu", hour=8, minute=0)
        _scheduler.start()
        print("[Scheduler] Gmail auto-sync aktiv: man+tor 08:00")
    yield
    if _scheduler.running:
        _scheduler.shutdown()

app = FastAPI(title="Organic Market Dashboard", lifespan=lifespan)

# Tillad upload op til 10 MB (mobilbilleder komprimeres i frontend til ~1.5 MB)
from starlette.middleware.base import BaseHTTPMiddleware
class _UploadLimitMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        if request.headers.get("content-type","").startswith("multipart"):
            cl = int(request.headers.get("content-length", 0))
            if cl > 10 * 1024 * 1024:
                from starlette.responses import JSONResponse
                return JSONResponse({"ok": False, "fejl": "Billedet er for stort (max 10 MB) — prøv at tage et nyt foto"}, status_code=400)
        return await call_next(request)
app.add_middleware(_UploadLimitMiddleware)
templates = Jinja2Templates(directory="templates")

# CORS: tillad kald fra TGTG Store Portal (til browser-baseret sync)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://store.toogoodtogo.com",
        "https://om-dashboard-production-0f3a.up.railway.app",
    ],
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization"],
)

SECRET_KEY        = os.environ.get("SECRET_KEY",         "skift-mig-i-railway-variables")
WEBHOOK_SECRET    = os.environ.get("WEBHOOK_SECRET",     "OM-Greve-2026-Hemlig")
DASHBOARD_USER    = os.environ.get("DASHBOARD_USERNAME", "linda")
DASHBOARD_PASS    = os.environ.get("DASHBOARD_PASSWORD", "")

signer = URLSafeTimedSerializer(SECRET_KEY)
SESSION_MAX_AGE   = 60 * 60 * 24 * 7  # 7 dage


# ── SESSION ───────────────────────────────────────────────────────────────────

def get_session(request: Request):
    token = request.cookies.get("session")
    if not token:
        return None
    try:
        return signer.loads(token, max_age=SESSION_MAX_AGE)
    except (BadSignature, SignatureExpired):
        return None


# ── SIDER ─────────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def root(request: Request):
    if not get_session(request):
        return RedirectResponse("/login", status_code=302)
    return RedirectResponse("/dashboard", status_code=302)


@app.get("/login", response_class=HTMLResponse)
async def login_side(request: Request):
    if get_session(request):
        return RedirectResponse("/dashboard", status_code=302)
    return templates.TemplateResponse("login.html", {"request": request, "fejl": None})


@app.post("/login", response_class=HTMLResponse)
async def login_post(
    request: Request,
    brugernavn: str = Form(...),
    adgangskode: str = Form(...),
):
    if brugernavn == DASHBOARD_USER and adgangskode == DASHBOARD_PASS:
        token = signer.dumps({"brugernavn": brugernavn})
        svar = RedirectResponse("/dashboard", status_code=303)
        svar.set_cookie("session", token, httponly=True, samesite="lax", max_age=SESSION_MAX_AGE)
        return svar
    return templates.TemplateResponse(
        "login.html",
        {"request": request, "fejl": "Forkert brugernavn eller adgangskode"},
    )


@app.get("/logout")
async def logout():
    svar = RedirectResponse("/login", status_code=302)
    svar.delete_cookie("session")
    return svar


@app.get("/dashboard", response_class=HTMLResponse)
async def dashboard(request: Request):
    if not get_session(request):
        return RedirectResponse("/login", status_code=302)
    from fastapi.responses import HTMLResponse as _HR
    resp = templates.TemplateResponse("dashboard.html", {"request": request})
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    return resp


# ── API ───────────────────────────────────────────────────────────────────────

def _kræv_login(request: Request):
    if not get_session(request):
        raise HTTPException(status_code=401, detail="Ikke logget ind")


@app.get("/api/data")
async def api_data(request: Request):
    _kræv_login(request)
    return database.hent_dashboard_data()


@app.get("/api/kpi")
async def api_kpi(request: Request, aar: Optional[int] = None):
    _kræv_login(request)
    return database.hent_kpi(aar)


@app.get("/api/salg/dag-db-detalje")
async def api_dag_db_detalje(request: Request):
    _kræv_login(request)
    return database.hent_dag_db_detalje()


@app.get("/api/debug/varenummer")
async def api_debug_varenummer(request: Request, navn: str = ""):
    _kræv_login(request)
    with database._conn() as conn:
        trans = conn.execute("""
            SELECT DISTINCT varenavn, varenummer, kategori,
                   COUNT(*) as linjer, MAX(dato) as seneste
            FROM transaktioner
            WHERE LOWER(varenavn) LIKE ?
            GROUP BY varenavn, varenummer, kategori
            ORDER BY seneste DESC LIMIT 20
        """, (f"%{navn.lower()}%",)).fetchall()
        stamdata = conn.execute("""
            SELECT id, sku, varenavn, type, pris_ex_moms, portioner
            FROM varestamdata
            WHERE LOWER(varenavn) LIKE ? OR LOWER(sku) LIKE ?
        """, (f"%{navn.lower()}%", f"%{navn.lower()}%")).fetchall()
        view = conn.execute("""
            SELECT varenavn, varenummer, antal, omsætning,
                   vf_korrekt, db_korrekt, dato
            FROM v_transaktioner
            WHERE LOWER(varenavn) LIKE ? AND dato = (SELECT MAX(dato) FROM transaktioner)
            LIMIT 5
        """, (f"%{navn.lower()}%",)).fetchall()
    return {
        "transaktioner": [dict(r) for r in trans],
        "varestamdata": [dict(r) for r in stamdata],
        "view_sample": [dict(r) for r in view]
    }


@app.get("/api/salg/idag")
async def api_idag(request: Request, aar: Optional[int] = None):
    _kræv_login(request)
    return database.hent_dag_produkter(aar)


@app.get("/api/salg/dag/{dato}")
async def api_dag_specificeret(request: Request, dato: str, aar: Optional[int] = None):
    _kræv_login(request)
    return database.hent_dag_produkter_by_date(dato, aar)


@app.get("/api/salg/dage")
async def api_dage(request: Request, n: int = 14, aar: Optional[int] = None):
    _kræv_login(request)
    return database.hent_dage(min(n, 365), aar)


@app.get("/api/salg/uger")
async def api_uger(request: Request, aar: Optional[int] = None):
    _kræv_login(request)
    return database.hent_uger(aar)


@app.get("/api/salg/timer")
async def api_timer(request: Request, aar: Optional[int] = None):
    _kræv_login(request)
    return database.hent_timer_idag(aar)


@app.get("/api/salg/timer/forrige-uge")
async def api_timer_forrige_uge(request: Request, aar: Optional[int] = None):
    _kræv_login(request)
    return database.hent_timer_forrige_uge(aar)


@app.get("/api/salg/timer/snit")
async def api_timer_snit(request: Request, aar: Optional[int] = None):
    _kræv_login(request)
    return database.hent_timer_snit(aar)


@app.get("/api/salg/kategorier")
async def api_kategorier(request: Request, aar: Optional[int] = None):
    _kræv_login(request)
    return database.hent_kategorier(aar)


@app.get("/api/salg/kategorier/uge")
async def api_kategorier_uge(request: Request, aar: Optional[int] = None):
    _kræv_login(request)
    return database.hent_kategorier_uge(aar)


@app.get("/api/salg/top")
async def api_top(request: Request, n: int = 20, aar: Optional[int] = None):
    _kræv_login(request)
    return database.hent_top_produkter(min(n, 100), aar)


@app.get("/api/salg/margin-analyse")
async def api_margin_analyse(request: Request, aar: Optional[int] = None, kategori: Optional[str] = None):
    _kræv_login(request)
    data = database.hent_margin_analyse(aar, kategori)
    # DEBUG: hvis ingen data, returnér info om hvad der er i databasen
    if not data:
        import sqlite3
        try:
            conn = sqlite3.connect(database.DB_PATH)
            trans_count = conn.execute("SELECT COUNT(*) FROM transaktioner").fetchone()[0]
            v_trans_count = conn.execute("SELECT COUNT(*) FROM v_transaktioner").fetchone()[0]
            conn.close()
            return {"debug": f"transaktioner={trans_count}, v_transaktioner={v_trans_count}", "data": []}
        except:
            pass
    return data


@app.get("/api/salg/aarsdata")
async def api_aarsdata(request: Request, aar: Optional[int] = None):
    _kræv_login(request)
    return database.hent_aarsdata(aar)


@app.get("/api/salg/trend")
async def api_trend(request: Request, dage: int = 21, aar: Optional[int] = None):
    _kræv_login(request)
    return database.hent_trend_analyse(min(dage, 90), aar)


@app.get("/api/salg/kaffe")
async def api_kaffe(request: Request, aar: Optional[int] = None):
    _kræv_login(request)
    return database.hent_kaffe_analyse(aar)


@app.get("/api/salg/dage-detaljer")
async def api_dage_detaljer(request: Request, n: int = 8, aar: Optional[int] = None):
    _kræv_login(request)
    return database.hent_dage_detaljer(min(n, 30), aar)


@app.get("/api/rapport-status")
async def rapport_status():
    info = database.hent_seneste_snapshot_info()
    return {
        "ok": True,
        "seneste_rapport": info.get("rapport_dato") if info else None,
        "indlæst":         info.get("indlæst_dato") if info else None,
    }


@app.get("/api/debug-ping")
async def debug_ping(request: Request):
    """Diagnostik: tjek session + database uden login-krav på selve ping."""
    has_session = bool(get_session(request))
    try:
        info = database.hent_seneste_snapshot_info()
        db_ok = True
        db_dato = info.get("rapport_dato") if info else None
    except Exception as e:
        db_ok = False
        db_dato = str(e)
    return {
        "session": has_session,
        "db_ok": db_ok,
        "db_dato": db_dato,
        "cookie_names": list(request.cookies.keys()),
    }


# ── WEBHOOK ───────────────────────────────────────────────────────────────────

@app.get("/api/bestilling/uger")
async def api_bestilling_uger(request: Request, aar: Optional[int] = None):
    _kræv_login(request)
    return database.hent_bestilling_uger(aar)


@app.get("/api/bestilling/uge/{uge}")
async def api_bestilling_uge(request: Request, uge: int, aar: Optional[int] = None):
    _kræv_login(request)
    if aar is None:
        from datetime import datetime
        aar = datetime.now().year
    return database.hent_bestilling_uge(uge, aar)



@app.post("/api/bestilling/opdater")
async def bestilling_opdater(request: Request):
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Ugyldig JSON")

    header_secret = request.headers.get("X-Webhook-Secret", "")
    if header_secret != WEBHOOK_SECRET and body.get("secret") != WEBHOOK_SECRET:
        raise HTTPException(status_code=401, detail="Ugyldig webhook secret")

    uge    = body.get("uge")
    aar    = body.get("aar")
    linjer = body.get("linjer", [])

    if not uge or not aar:
        raise HTTPException(status_code=400, detail="Mangler uge eller aar")
    if not linjer:
        raise HTTPException(status_code=400, detail="Ingen bestillingslinjer")

    antal = database.gem_ugebestilling(int(uge), int(aar), linjer)

    # Auto-opdater stamdata fra bestillingslinjer (SKU + pris_ex_moms)
    stamdata_linjer = [
        {"sku": l["varenummer"], "varenavn": l["varenavn"],
         "type": "Bagværk", "pris_ex_moms": l["pris_ex_moms"]}
        for l in linjer
        if l.get("varenummer") and l.get("pris_ex_moms", 0) > 0
    ]
    if stamdata_linjer:
        database.gem_stamdata_bulk(stamdata_linjer)

    return {"ok": True, "uge": uge, "aar": aar, "linjer": antal}


@app.get("/api/bagvaerk/dag/{uge}")
async def api_bagvaerk_dag(request: Request, uge: int, aar: Optional[int] = None):
    _kræv_login(request)
    if aar is None:
        from datetime import date
        aar = date.today().year
    return database.hent_bagvaerk_dag_sammenligning(uge, aar)


@app.get("/api/bager/svind")
async def api_bager_svind(request: Request, aar: Optional[int] = None):
    _kræv_login(request)
    return database.hent_svind_data(aar)


@app.post("/api/bestilling/gem-manuel")
async def bestilling_gem_manuel(request: Request):
    _kræv_login(request)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Ugyldig JSON")
    uge  = body.get("uge")
    aar  = body.get("aar")
    vn   = body.get("varenummer")
    dag  = body.get("dag")
    antal = body.get("antal")
    if uge is None or aar is None or vn is None or dag is None or antal is None:
        raise HTTPException(status_code=400, detail="Mangler felter")
    database.gem_bestilling_manuel(int(uge), int(aar), str(vn), str(dag), int(antal))
    return {"ok": True}


@app.post("/api/bestilling/vurder")
async def api_beregner_vurder(request: Request):
    """AI vurderer bestillingen og returnerer tekst + strukturerede justeringer."""
    _kræv_login(request)
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return {"ok": False, "fejl": "⚠️ AI-vurdering ikke tilgængelig — ANTHROPIC_API_KEY ikke sat på Railway"}
    try:
        body = await request.json()
        import anthropic as _ant, json as _json

        # Debug: check hvis API-nøgle bliver læst
        if not api_key or api_key.strip() == "":
            print("[ERROR] ANTHROPIC_API_KEY er tom eller ikke sat!")
            return {"ok": False, "fejl": "API-nøgle er tom. Kontakt administrator."}
        print(f"[DEBUG] API-nøgle længde: {len(api_key)}, starter med: {api_key[:10]}...")

        # Validér at vi har de vigtigste felter
        required_fields = ['uge', 'aar', 'dag_totaler', 'produkter']
        for field in required_fields:
            if field not in body or body[field] is None:
                return {"ok": False, "fejl": f"Manglende felt: {field}"}

        # Sanitér input-data: fjern problematiske tegn der kan bryde Claude's JSON output
        def sanitize_prompt_input(text):
            if not isinstance(text, str):
                return str(text) if text is not None else ""
            # Erstat newlines med space
            text = text.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')
            # Begræns flere mellemrum
            import re
            text = re.sub(r'  +', ' ', text)
            return text.strip()

        # Brug .format() i stedet for f-string for at undgå curly-brace fortolkning
        prompt = """Du er bestillingsrådgiver for Organic Market Greve — specialbutik med bageri.

═══ FORRETNINGSLOGIK — FORSTÅ DETTE FØR ALT ANDET ═══
Vores to LIGE STORE risici er:
1. FOR MEGET på svage dage → retur/TGTG → direkte tab (kostpris + arbejdstid)
2. FOR LIDT på stærke dage → tomme hylder → tabt salg og skuffede kunder

Begge er tabsbringende. Det handler om DAG-PRÆCISION — ikke bare ugetotal.
TGTG-mål: under 800 kr/uge. Over 1.200 kr = vi overbestiller på svage dage.
Lørdage og fredage er typisk stærke. Mandage og tirsdage typisk svage.
Begivenheder kan VENDE dette mønster helt.

⚠ DATAKVALITET — disse forbehold gælder ALTID:
• Shopbox er manuelt tastet → varenavn/antal kan have fejl → sell-through undervurderer reelt salg
• MobilePay-salg er ikke varekoblet → en del af bagværkssalget mangler i produkt-tallene
• Brug derfor TGTG-kr og retur-stk som primære signaler — de er mere præcise end Shopbox-antal
• Vær forsigtig med store reduktioner baseret på lav sell-through alene
═══════════════════════════════════════════════════════

BESTILLINGSUGE {}/{} ({}):
Begivenhed: {}
Sæsonindeks: {} · Vækst: {}
TGTG seneste uge: {}
Dag-snit fra historik: {}

HISTORISK SELL-THROUGH (solgt/bestilt % pr. kategori pr. dag — seneste 10 uger):
{}
[>95% = sandsynligvis udsolgt · <75% = spild/TGTG-risiko]
DATAKVALITET: {}
(Shopbox er manuelt tastet — sell-through % undervurderer reelt salg)

DAGSTOTALER DENNE BESTILLING:
{}

PRODUKTER PR. DAG:
{}

Vurder BEGGE risici for HVER dag:
- Er stærke dage (fre, lør, begivenhedsdage) bestilt højt nok? Tomme hylder = tabt salg.
- Er svage dage bestilt for højt? Overskud = TGTG-tab.

Svar UDELUKKENDE med JSON. Intet andet tekst før eller efter. Ingen markdown, ingen backticks, BARE JSON.

SVAR-FORMAT:
{{
  "vurdering": "kort vurdering her",
  "klar": true,
  "justeringer": []
}}

INSTRUKTIONER:
- vurdering: 1-2 sætninger. VIGTIG: Kun disse tegn tilladt i tekst: bogstaver, tal, mellemrum, punktum, komma, bindesteg
- klar: true hvis bestilling virker fin, false hvis problemer
- justeringer: liste af justeringer. HVIS INGEN justeringer, brug tom liste []
- Hver justering: varenavn (tekst), dag (man/tir/ons/tor/fre/loe/son), fra (tal), til (tal), grund (kort tekst, kun bogstaver/tal/mellemrum)
- grund eksempler: tabt salg, hoj TGTG risiko, lav sell-through, for meget retur
- Maksimum 10 justeringer
- VIGTIG REGEL: Ingen citationstegn, apostrof, eller special tegn i tekstværdier. Kun bogstaver a-z, tal 0-9, mellemrum, punktum, komma, bindesteg""".format(
            sanitize_prompt_input(body.get('uge')),
            sanitize_prompt_input(body.get('aar')),
            sanitize_prompt_input(body.get('dato_range','')),
            sanitize_prompt_input(body.get('event','ingen')),
            sanitize_prompt_input(body.get('si',1.0)),
            sanitize_prompt_input(body.get('vaekst','?')),
            sanitize_prompt_input(body.get('tgtg','ingen data')),
            sanitize_prompt_input(body.get('dag_snit','')),
            sanitize_prompt_input(body.get('sellthrough', 'ingen data')),
            sanitize_prompt_input(body.get('mobilepay_andel', 'MobilePay-andel ukendt')),
            sanitize_prompt_input(body.get('dag_totaler','')),
            sanitize_prompt_input(body.get('produkter',''))
        )

        client = _ant.Anthropic(api_key=api_key)
        msg = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=800,
            messages=[{"role": "user", "content": prompt}]
        )
        raw = msg.content[0].text.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.lower().startswith("json"): raw = raw[4:]
        raw = raw.strip()

        # Log Claude's raw response for debugging
        print(f"[DEBUG] Claude response length: {len(raw)}, first 300 chars:\n{raw[:300]}")

        import re

        def parse_claude_json(text):
            """Forsøg at parse Claude's JSON response med flere strategier"""
            # Strategi 1: Direkte parsing
            try:
                return _json.loads(text)
            except _json.JSONDecodeError as e1:
                pass

            # Strategi 2: Fjern alle newlines og normaliser whitespace
            try:
                text_fixed = text.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')
                text_fixed = re.sub(r'  +', ' ', text_fixed)
                return _json.loads(text_fixed)
            except _json.JSONDecodeError as e2:
                pass

            # Strategi 3: Fjern markdown-wrapper
            try:
                text_fixed = text.replace('```json', '').replace('```', '')
                text_fixed = text_fixed.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')
                text_fixed = re.sub(r'  +', ' ', text_fixed)
                return _json.loads(text_fixed)
            except _json.JSONDecodeError as e3:
                pass

            # Strategi 4: Escape alle uescapede citationstegn inden for strenge
            try:
                # Find alle { og } for at lokalisere strings
                in_string = False
                escaped = False
                result = []
                for i, char in enumerate(text):
                    if char == '\\' and not escaped:
                        escaped = True
                        result.append(char)
                    elif char == '"' and not escaped:
                        in_string = not in_string
                        result.append(char)
                    elif char in '\n\r' and in_string:
                        result.append(' ')
                    else:
                        escaped = False
                        result.append(char)

                text_fixed = ''.join(result)
                return _json.loads(text_fixed)
            except _json.JSONDecodeError as e4:
                pass

            # Hvis alle strategier fejler, kast fejl med info
            raise ValueError(f"Kunne ikke parse JSON. Sidste fejl: {e4.msg if 'e4' in locals() else 'unknown'}")

        try:
            parsed = parse_claude_json(raw)
        except Exception as je:
            print(f"[DEBUG] JSON parse fejl: {str(je)}")
            print(f"[DEBUG] Raw Claude response: {raw[:500]}")
            return {"ok": False, "fejl": f"JSON parse fejl: {str(je)[:100]}"}  # Simplificeret fejlbesked

        return {
            "ok": True,
            "vurdering":   parsed.get("vurdering", ""),
            "klar":        parsed.get("klar", True),
            "justeringer": parsed.get("justeringer", []),
        }
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        # Log den faktiske fejl
        print(f"Fejl i bestilling/vurder: {type(e).__name__}: {str(e)}\n{tb}")
        return {"ok": False, "fejl": f"Bestillingsværdier kunne ikke genereres: {type(e).__name__} - {str(e)[:50]}"}


@app.get("/api/bestilling/sellthrough")
async def api_sellthrough(request: Request, uger: int = 10):
    _kræv_login(request)
    return database.hent_sellthrough_analyse(uger)


@app.get("/api/bestilling/vejr-debug")
async def api_vejr_debug(request: Request, uge: int, aar: int):
    """Debug: vis præcis hvilke vejrdata der bruges til AI for en given uge."""
    _kræv_login(request)
    from datetime import date, timedelta
    vejr = database.hent_vejr_forecast()
    fc   = vejr.get("forecast", {})
    mon  = date.fromisocalendar(aar, uge, 1)
    DAG  = ["Man","Tir","Ons","Tor","Fre","Lør","Søn"]
    result = []
    for i in range(7):
        dag = mon + timedelta(days=i)
        ds  = dag.isoformat()
        v   = fc.get(ds)
        result.append({
            "dag": DAG[i], "dato": ds,
            "fundet": v is not None,
            "data": v
        })
    return {"uge": uge, "aar": aar, "mandag": mon.isoformat(),
            "forecast_keys_sample": list(fc.keys())[:5], "dage": result}


@app.post("/api/bestilling/kontekst")
async def api_beregner_kontekst(request: Request):
    _kræv_login(request)
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return {"ok": False, "fejl": "ANTHROPIC_API_KEY ikke konfigureret"}
    try:
        body = await request.json()
        uge  = int(body.get("uge", 0))
        aar  = int(body.get("aar", 0))
        dag_totaler = body.get("dag_totaler", {})
        produkter   = body.get("produkter", [])
        vejr_js     = body.get("vejr", {})
        # Brug JS vejr-cache (fra strip) hvis den har data for den relevante uge
        # Ellers fallback til server-cache — begge kan have forskellig alder
        from datetime import date, timedelta
        mon_check = date.fromisocalendar(uge, uge, 1) if False else date.fromisocalendar(aar, uge, 1)
        mon_str   = mon_check.isoformat()
        js_har_data = bool(vejr_js.get("forecast", {}).get(mon_str))
        vejr = vejr_js if js_har_data else database.hent_vejr_forecast()
        return database.generer_beregner_kontekst(uge, aar, api_key, dag_totaler, produkter, vejr)
    except Exception as e:
        return {"ok": False, "fejl": str(e)}


@app.get("/api/bestilling/anbefaling")
async def api_bestillings_anbefaling(
    request: Request,
    uge: Optional[int] = None,
    aar: Optional[int] = None,
):
    _kræv_login(request)
    if uge is None:
        from datetime import date
        iso = date.today().isocalendar()
        uge = iso[1] + 1
        aar = iso[0]
        if uge > 52:
            uge = 1
            aar += 1
    if aar is None:
        from datetime import date
        aar = date.today().year
    try:
        return database.hent_bestillings_uge(int(uge), int(aar))
    except Exception as e:
        import traceback
        error_msg = str(e)
        print(f"ERROR in hent_bestillings_uge({uge}, {aar}): {error_msg}")
        traceback.print_exc()
        return {
            "error": f"Fejl ved bestillingsberegning: {error_msg}",
            "uge": uge,
            "aar": aar
        }


@app.post("/api/bestilling/management-analyse")
async def api_management_analyse(
    request: Request,
    uge: int,
    aar: int,
):
    """Management-analyse med AKTUELLE værdier fra frontend (ikke database)."""
    _kræv_login(request)
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return {"analyse": None, "fejl": "ANTHROPIC_API_KEY ikke konfigureret i Railway"}

    try:
        body = await request.json()
    except Exception:
        body = {}

    # Hent basis-info fra database for sammenligning
    d = database.hent_bestillings_uge(int(uge), int(aar))
    if "fejl" in d or "error" in d:
        return {"analyse": None, "fejl": d.get("fejl") or d.get("error")}

    # Hvis frontend sender produkter, brug dem. Ellers fall back til database
    produkter_data = body.get("produkter", [])

    if produkter_data:
        # Brug AKTUELLE værdier fra frontend
        total_stk = sum(p.get("total", 0) for p in produkter_data)
        # Estimat af pris baseret på basis-data (vi har ikke detaljerede priser fra frontend)
        total_kr = total_stk * 12  # Approks gennemsnit

        # Byg kategori-info fra basis for sammenligning
        kat_map = {}
        for p in d.get("produkter", []):
            kat = p.get("kategori") or "Øvrige"
            basis_total = sum((p.get("basis") or {}).values())
            if kat not in kat_map:
                kat_map[kat] = {"basis": 0}
            kat_map[kat]["basis"] += basis_total

        # Tilføj aktuelle værdier
        for p_data in produkter_data:
            navn = p_data.get("navn", "")
            # Match kategori fra basis-data
            kat = "Øvrige"
            for p_basis in d.get("produkter", []):
                if p_basis.get("varenavn", "").strip() == navn.strip():
                    kat = p_basis.get("kategori", "Øvrige")
                    break

            if kat not in kat_map:
                kat_map[kat] = {"basis": 0, "aktuel": 0}
            if "aktuel" not in kat_map[kat]:
                kat_map[kat]["aktuel"] = 0
            kat_map[kat]["aktuel"] += p_data.get("total", 0)
    else:
        # Fall back til database værdier
        total_stk = d.get("total_stk", 0)
        total_kr = d.get("total_kr", 0)

        kat_map = {}
        for p in d.get("produkter", []):
            kat = p.get("kategori") or "Øvrige"
            anbefalet = p.get("total_anbefalet", 0)
            basis_total = sum((p.get("basis") or {}).values())
            if kat not in kat_map:
                kat_map[kat] = {"aktuel": 0, "basis": 0}
            kat_map[kat]["aktuel"] += anbefalet
            kat_map[kat]["basis"] += basis_total

    kat_linjer = []
    for kat, v in sorted(kat_map.items(), key=lambda x: -x[1].get("aktuel", x[1].get("anbefalet", 0))):
        aktuel = v.get("aktuel", v.get("anbefalet", 0))
        basis = v.get("basis", 0)
        diff = aktuel - basis
        pct = round(diff / basis * 100, 1) if basis > 0 else 0
        kat_linjer.append(
            f"  {kat}: {round(aktuel)} stk (basis: {round(basis)} stk, "
            f"ændring: {'+' if diff >= 0 else ''}{round(diff)} stk / {'+' if pct >= 0 else ''}{pct}%)"
        )

    evt = d.get("event")
    evt_txt = f"\nBegivenhed denne uge: {evt['navn']} — {evt.get('note','')} (faktor ×{evt.get('factor',1)})" if evt else "\nIngen registrerede begivenheder denne uge."

    prompt = f"""Du er den erfarne bageri-chef og management-rådgiver for Organic Market i Greve — et dansk specialbageri.

Gennemgå nedenstående bestillingsdata for uge {d['maal_uge']} {d['maal_aar']} og giv en kort, konkret management-vurdering på dansk.

BESTILLINGSDATA (AKTUELLE VÆRDIER):
- Måluge: {d['maal_uge']} {d['maal_aar']} ({d.get('dato_range','')})
- Basisuge: {d['basis_uge']} {d['basis_aar']}
- Forslået total: {round(total_stk)} stk / {round(total_kr)} kr ex moms
- Sæsonindeks (SI): {d['si']:.2f} ({'+' if d['si'] >= 1 else ''}{round((d['si']-1)*100)}% ift. neutral)
- Væksttrend: {'+' if d['vaekst_pct'] >= 0 else ''}{d['vaekst_pct']:.1f}%
- Too Good To Go: {round(d['tgtg_kr'])} kr/uge {'⚠ FOR HØJ' if d.get('tgtg_advarsel') else '✓ OK' if d.get('tgtg_ok') else ''}
{evt_txt}

KATEGORI-FORDELING (aktuel vs. basis):
{chr(10).join(kat_linjer)}

Giv en management-vurdering med:
1. En samlet konklusion (2-3 sætninger): ser bestillingen fornuftig ud?
2. Hvad har ændret sig vs. basisugen og hvorfor (sæson, trend, begivenhed)?
3. 1-2 konkrete råd eller ting at holde øje med.

Vær direkte og konkret. Brug tal. Maks 200 ord."""

    try:
        import anthropic as _ant
        client = _ant.Anthropic(api_key=api_key)
        msg = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=600,
            messages=[{"role": "user", "content": prompt}],
        )
        tekst = msg.content[0].text
        return {"analyse": tekst}
    except Exception as e:
        print(f"[ERROR] management-analyse fejl: {e}")
        import traceback
        traceback.print_exc()
        return {"analyse": None, "fejl": str(e)}


@app.get("/api/bestilling/eksport")
async def api_bestilling_eksport(
    request: Request,
    uge: Optional[int] = None,
    aar: Optional[int] = None,
):
    _kræv_login(request)
    if uge is None:
        from datetime import date
        iso = date.today().isocalendar()
        uge = iso[1] + 1
        aar = iso[0]
        if uge > 52:
            uge = 1
            aar += 1
    if aar is None:
        from datetime import date
        aar = date.today().year

    d = database.hent_bestillings_uge(int(uge), int(aar))
    if "fejl" in d:
        raise HTTPException(status_code=404, detail=d["fejl"])

    xlsx_bytes = _byg_bestilling_xlsx(d)
    filename = f"Bestilling uge {d['maal_uge']} {d['maal_aar']}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _byg_bestilling_xlsx(d: dict) -> bytes:
    from openpyxl import load_workbook

    DAGE = ['man', 'tir', 'ons', 'tor', 'fre', 'loe', 'son']
    TEMPLATE = Path(__file__).parent / "bestilling_template.xlsx"

    wb = load_workbook(str(TEMPLATE))
    ws = wb.active

    # Opdater ugenummer i A2
    ws['A2'] = f"Uge {d['maal_uge']}"

    # Byg SKU → produkt-map fra bestillingsdata
    prod_map: dict = {}
    for p in d.get("produkter", []):
        sku = p.get("varenummer")
        if sku:
            try:
                prod_map[int(str(sku).strip())] = p
            except (ValueError, TypeError):
                pass

    # Find "I alt"-rækker (col C = 'I alt') — dem rører vi ikke E-K på
    ialt_rows: set = set()
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=False):
        if row[2].value == 'I alt':
            ialt_rows.add(row[0].row)

    # Ryd E-K i alle data-rækker og udfyld fra bestillingsdata via SKU-match
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=False):
        rn = row[0].row
        if rn in ialt_rows:
            continue

        # Ryd dagkolonner E-K (tuple-index 4–10)
        for i in range(4, 11):
            row[i].value = None

        sku_val = row[1].value  # kolonne B = varenummer
        if sku_val and isinstance(sku_val, (int, float)):
            sku_int = int(sku_val)
            if sku_int in prod_map:
                p = prod_map[sku_int]
                # Opdater pris (kolonne D = index 3)
                if p.get("pris_ex_moms"):
                    row[3].value = p["pris_ex_moms"]
                # Indsæt anbefalede antal (E-K)
                anb = p.get("anbefalet", {})
                for i, dag in enumerate(DAGE):
                    qty = anb.get(dag, 0)
                    row[4 + i].value = qty if qty else None

    # Rad 68 har hardkodede dagssummer for sektion 3 (Kage/muffin-rækker 63–67)
    # Genberegn disse fra de opdaterede produktrækker
    if 68 not in ialt_rows:  # sikkerhed — det er en I alt-række, men E-K er tal, ikke formler
        dag_sums = [0] * 7
        for r in range(63, 68):
            data_row = list(ws.iter_rows(min_row=r, max_row=r, values_only=False))[0]
            for i in range(7):
                v = data_row[4 + i].value
                if v and isinstance(v, (int, float)):
                    dag_sums[i] += int(v)
        row68 = list(ws.iter_rows(min_row=68, max_row=68, values_only=False))[0]
        for i in range(7):
            row68[4 + i].value = dag_sums[i] if dag_sums[i] else None

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── BASIS BESTILLING API ──────────────────────────────────────────────────────

@app.get("/api/basis-bestilling/")
async def api_basis_bestilling_get(request: Request):
    """Hent alle basis-bestillinger."""
    _kræv_login(request)
    return {"ok": True, "data": database.hent_basis_bestilling()}


@app.get("/api/basis-bestilling/dag/{dag}")
async def api_basis_bestilling_dag(request: Request, dag: str):
    """Hent basis-bestillinger for en specifik dag."""
    _kræv_login(request)
    if dag not in ['man', 'tir', 'ons', 'tor', 'fre', 'loe', 'son']:
        raise HTTPException(status_code=400, detail="Ugyldigt dagnavn")
    return {"ok": True, "data": database.hent_basis_bestilling_ved_dag(dag)}


@app.post("/api/basis-bestilling/gem")
async def api_basis_bestilling_gem(request: Request):
    """Gem eller opdater en basis-bestillingslinje."""
    _kræv_login(request)
    try:
        body = await request.json()
        varenummer = body.get('varenummer')
        varenavn = body.get('varenavn', '')
        dag = body.get('dag')
        antal = int(body.get('anbefalet_antal', 0))
        kategori = body.get('kategori', '')

        if not varenummer or not dag:
            raise ValueError("varenummer og dag er påkrævet")
        if dag not in ['man', 'tir', 'ons', 'tor', 'fre', 'loe', 'son']:
            raise ValueError(f"Ugyldigt dagnavn: {dag}")

        database.gem_basis_bestilling(varenummer, varenavn, dag, antal, kategori)
        return {"ok": True, "varenummer": varenummer, "dag": dag, "anbefalet_antal": antal}
    except Exception as e:
        return {"ok": False, "fejl": str(e)}


@app.post("/api/basis-bestilling/bulk-gem")
async def api_basis_bestilling_bulk_gem(request: Request):
    """Batch-gem flere basis-bestillinger."""
    _kræv_login(request)
    try:
        body = await request.json()
        updates = body.get('updates', [])
        if not isinstance(updates, list):
            raise ValueError("updates skal være en liste")

        database.bulk_opdater_basis_bestilling(updates)
        return {"ok": True, "updated_count": len(updates)}
    except Exception as e:
        return {"ok": False, "fejl": str(e)}


@app.post("/api/basis-bestilling/slet")
async def api_basis_bestilling_slet(request: Request):
    """Slet en basis-bestillingslinje."""
    _kræv_login(request)
    try:
        body = await request.json()
        varenummer = body.get('varenummer')
        dag = body.get('dag')

        if not varenummer or not dag:
            raise ValueError("varenummer og dag er påkrævet")

        database.slet_basis_bestilling_linje(varenummer, dag)
        return {"ok": True, "varenummer": varenummer, "dag": dag}
    except Exception as e:
        return {"ok": False, "fejl": str(e)}


@app.get("/api/basis-bestilling/produkter")
async def api_basis_bestilling_produkter(request: Request):
    """Hent alle produkter i basis_bestilling."""
    _kræv_login(request)
    return {"ok": True, "data": database.hent_basis_bestilling_produkter()}


@app.get("/api/helligdage")
async def api_helligdage(request: Request, aar: Optional[int] = None):
    """Hent helligdage."""
    _kræv_login(request)
    return {"ok": True, "data": database.hent_helligdage(aar)}


@app.post("/api/bager/upload-pdf")
async def bager_upload_pdf(request: Request, fil: UploadFile = File(...)):
    """Parse bager-faktura PDF med Claude og returner ekstraherede felter."""
    _kræv_login(request)
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return {"ok": False, "fejl": "ANTHROPIC_API_KEY ikke konfigureret"}
    try:
        pdf_bytes = await fil.read()
        pdf_b64   = __import__("base64").b64encode(pdf_bytes).decode()
        import anthropic as _ant
        client = _ant.Anthropic(api_key=api_key)
        msg = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=512,
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "document",
                        "source": {
                            "type": "base64",
                            "media_type": "application/pdf",
                            "data": pdf_b64,
                        },
                    },
                    {
                        "type": "text",
                        "text": (
                            "Dette er en ugentlig bageri-faktura til Organic Market Greve.\n"
                            "Ekstraher præcist følgende felter og returner KUN valid JSON (ingen forklaring):\n"
                            "{\n"
                            '  "uge": <ugenummer som heltal>,\n'
                            '  "aar": <årstal som heltal>,\n'
                            '  "retur_wiener": <returneret wienerbrød antal stk, 0 hvis ikke nævnt>,\n'
                            '  "retur_boller": <returnerede boller antal stk, 0 hvis ikke nævnt>,\n'
                            '  "tgtg": <Too Good To Go antal stk, 0 hvis ikke nævnt>,\n'
                            '  "b_kvali": <kvalitetskreditering beløb i kr (positivt tal), 0 hvis ikke nævnt>,\n'
                            '  "retur_ialt": <total returkredit i kr (positivt tal), 0 hvis ikke nævnt>,\n'
                            '  "faktura": <faktura total at betale i kr, 0 hvis ikke nævnt>\n'
                            "}"
                        ),
                    },
                ],
            }],
        )
        import json as _json
        raw = msg.content[0].text.strip()
        # Trim markdown code fences hvis til stede
        raw = raw.strip("`")
        if raw.lower().startswith("json"):
            raw = raw[4:].strip()
        data = _json.loads(raw)
        # Valider felter
        for f in ("uge", "aar", "retur_wiener", "retur_boller", "tgtg", "b_kvali", "retur_ialt", "faktura"):
            if f not in data:
                data[f] = 0
        return {"ok": True, "data": data}
    except Exception as exc:
        return {"ok": False, "fejl": str(exc)}


@app.post("/api/retur/scan")
async def retur_scan(request: Request, fil: UploadFile = File(...)):
    """Upload retur-seddel foto → Claude Vision udtrækker produkter + antal."""
    _kræv_login(request)
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        return {"ok": False, "fejl": "ANTHROPIC_API_KEY ikke konfigureret"}
    img_bytes = await fil.read()
    img_b64 = base64.b64encode(img_bytes).decode()
    media_type = fil.content_type or "image/jpeg"
    if media_type not in ("image/jpeg", "image/png", "image/gif", "image/webp"):
        media_type = "image/jpeg"
    try:
        import anthropic as _ant
        import json as _json
        client = _ant.Anthropic(api_key=api_key)
        msg = client.messages.create(
            model="claude-opus-4-7",
            max_tokens=1024,
            messages=[{"role": "user", "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": img_b64}},
                {"type": "text", "text": (
                    "Dette er en retur-seddel fra et dansk bageri (Organic Market Greve).\n"
                    "Find alle produkter hvor der er skrevet et håndskrevet tal i kolonnen 'Retur antal'.\n"
                    "Ignorer produkter markeret 'krediteres ikke' eller 'tages ikke retur'.\n"
                    "Ignorer produkter med tom eller ingen retur-antal.\n\n"
                    "Kategoriser hvert produkt:\n"
                    "- 'boller': Hvedeboller, Müslibolle, Kernebolle, bolle-produkter\n"
                    "- 'wienerbroed': Croissant, Tebirkes, Kanel snegl, Wiener, Spandauer, Romsnegle, Kanelsnurre, Kardemomme, og alt andet wienerbrød\n\n"
                    "Returner KUN valid JSON uden forklaring:\n"
                    '{\"items\":[{\"produkt\":\"navn\",\"antal\":N,\"kategori\":\"boller|wienerbroed\"}]}'
                )},
            ]}],
        )
        raw = msg.content[0].text.strip().strip("`")
        if raw.lower().startswith("json"):
            raw = raw[4:].strip()
        data = _json.loads(raw)
        validated = [
            {"produkt": str(it["produkt"]), "antal": max(1, int(it["antal"])),
             "kategori": it.get("kategori", "wienerbroed") if it.get("kategori") in ("boller", "wienerbroed") else "wienerbroed"}
            for it in data.get("items", []) if it.get("produkt") and int(it.get("antal", 0)) > 0
        ]
        return {"ok": True, "items": validated}
    except Exception as e:
        return {"ok": False, "fejl": str(e)}


@app.post("/api/retur/gem")
async def retur_gem(request: Request):
    """Gem bekræftede retur-detaljer for en uge."""
    _kræv_login(request)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Ugyldig JSON")
    items = body.get("items", [])
    dato  = body.get("dato") or datetime.now().strftime("%Y-%m-%d")
    if not dato:
        raise HTTPException(status_code=400, detail="Mangler dato")
    # uge/aar er valgfrie — gem_retur_detaljer beregner dem fra datoen
    antal = database.gem_retur_detaljer(0, 0, items, dato)
    return {"ok": True, "antal": antal}


@app.get("/api/retur/status")
async def retur_status(request: Request):
    _kræv_login(request)
    return database.hent_retur_kpi()


@app.get("/api/retur/debug")
async def retur_debug(request: Request):
    _kræv_login(request)
    from datetime import date, timedelta
    today = date.today()
    weekday = today.weekday()
    this_monday = today - timedelta(days=weekday)
    prev_sunday = this_monday - timedelta(days=1)
    prev_iso = prev_sunday.isocalendar()
    aktuel_uge = int(prev_iso[1])
    aktuel_aar = int(prev_iso[0])
    with database._conn() as conn:
        # Hvad er i ugebestillinger for aktuel uge?
        alle = conn.execute("""
            SELECT u.varenavn, u.total_antal,
                   COALESCE(v.type,'—') AS type
            FROM ugebestillinger u
            LEFT JOIN varestamdata v ON LOWER(TRIM(u.varenavn)) = LOWER(TRIM(v.varenavn))
            WHERE u.uge=? AND u.aar=?
            ORDER BY v.type, u.varenavn
        """, (aktuel_uge, aktuel_aar)).fetchall()
        b = conn.execute("""
            SELECT COALESCE(SUM(total_antal),0) AS t FROM ugebestillinger
            WHERE uge=? AND aar=? AND LOWER(varenavn) LIKE '%bolle%'
        """, (aktuel_uge, aktuel_aar)).fetchone()
        w = conn.execute("""
            SELECT COALESCE(SUM(total_antal),0) AS t FROM ugebestillinger
            WHERE uge=? AND aar=? AND (
                LOWER(varenavn) LIKE '%croissant%' OR LOWER(varenavn) LIKE '%crossaint%' OR
                (LOWER(varenavn) LIKE '%birkes%' AND LOWER(varenavn) NOT LIKE '%hvede%') OR
                LOWER(varenavn) LIKE '%snegl%' OR
                LOWER(varenavn) LIKE '%snurrer%' OR LOWER(varenavn) LIKE '%snurr%' OR
                LOWER(varenavn) LIKE '%spandauer%' OR LOWER(varenavn) LIKE '%wienerstang%' OR
                LOWER(varenavn) LIKE '%kanelstang%' OR LOWER(varenavn) LIKE '%frøsnapper%'
            )
        """, (aktuel_uge, aktuel_aar)).fetchone()
        boller_varer = conn.execute("""
            SELECT varenavn, total_antal FROM ugebestillinger
            WHERE uge=? AND aar=? AND LOWER(varenavn) LIKE '%bolle%'
        """, (aktuel_uge, aktuel_aar)).fetchall()
        wiener_varer = conn.execute("""
            SELECT varenavn, total_antal FROM ugebestillinger
            WHERE uge=? AND aar=? AND (
                LOWER(varenavn) LIKE '%croissant%' OR LOWER(varenavn) LIKE '%crossaint%' OR
                (LOWER(varenavn) LIKE '%birkes%' AND LOWER(varenavn) NOT LIKE '%hvede%') OR
                LOWER(varenavn) LIKE '%snegl%' OR
                LOWER(varenavn) LIKE '%snurrer%' OR LOWER(varenavn) LIKE '%snurr%' OR
                LOWER(varenavn) LIKE '%spandauer%' OR LOWER(varenavn) LIKE '%wienerstang%' OR
                LOWER(varenavn) LIKE '%kanelstang%' OR LOWER(varenavn) LIKE '%frøsnapper%'
            )
        """, (aktuel_uge, aktuel_aar)).fetchall()
        retur = conn.execute("""
            SELECT kategori, SUM(antal) AS antal
            FROM retur_detaljer WHERE uge=? AND aar=?
            GROUP BY kategori
        """, (aktuel_uge, aktuel_aar)).fetchall()
    bestilt_b = round(b['t'] or 0)
    bestilt_w = round(w['t'] or 0)
    return {
        "dato_idag": str(today),
        "weekday": weekday,
        "aktuel_uge": aktuel_uge,
        "aktuel_aar": aktuel_aar,
        "bestilt_boller": bestilt_b,
        "bestilt_wiener": bestilt_w,
        "max_boller_10pct": round(bestilt_b * 0.10),
        "max_wiener_135pct": round(bestilt_w * 0.135),
        "retur_registreret": [dict(r) for r in retur],
        "boller_tæller_med": [dict(r) for r in boller_varer],
        "wiener_tæller_med": [dict(r) for r in wiener_varer],
        "alle_varer_i_uge": [dict(r) for r in alle],
    }


@app.get("/api/retur/uge/{uge}/{aar}")
async def retur_uge_data(request: Request, uge: int, aar: int):
    _kræv_login(request)
    return database.hent_retur_uge(uge, aar)


@app.get("/api/retur/historik")
async def retur_historik(request: Request, n: int = 60):
    _kræv_login(request)
    return database.hent_retur_historik(n)


@app.get("/api/retur/dag/{dato}")
async def retur_dag(request: Request, dato: str):
    _kræv_login(request)
    return database.hent_retur_dag(dato)


@app.post("/api/bager/retur-opdater")
async def bager_retur_opdater(request: Request):
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Ugyldig JSON")
    header_secret = request.headers.get("X-Webhook-Secret", "")
    if header_secret != WEBHOOK_SECRET and body.get("secret") != WEBHOOK_SECRET:
        raise HTTPException(status_code=401, detail="Ugyldig webhook secret")
    linjer = body.get("linjer", [])
    if not linjer:
        raise HTTPException(status_code=400, detail="Ingen linjer")
    antal = database.gem_bager_regnskab(linjer)
    return {"ok": True, "linjer": antal}


@app.get("/api/mobilepay")
async def api_mobilepay(request: Request):
    _kræv_login(request)
    return database.hent_mobilepay()


@app.post("/api/mobilepay/gem")
async def mobilepay_gem(request: Request):
    _kræv_login(request)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Ugyldig JSON")
    aar       = body.get("aar")
    maaned    = body.get("maaned")
    omsaetning = body.get("omsaetning")
    if aar is None or maaned is None or omsaetning is None:
        raise HTTPException(status_code=400, detail="Mangler felter")
    database.gem_mobilepay(int(aar), int(maaned), float(omsaetning))
    return {"ok": True}


@app.post("/api/mobilepay/dagssalg")
async def mobilepay_dagssalg(request: Request):
    """Webhook: modtager daglig MP-omsætning fra mobilepay_sync.py."""
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Ugyldig JSON")
    if body.get("secret") != WEBHOOK_SECRET:
        raise HTTPException(status_code=403, detail="Forkert secret")
    linjer = body.get("linjer", [])
    if not linjer:
        return {"ok": True, "linjer": 0}
    try:
        count = database.gem_mobilepay_dag(linjer)
        return {"ok": True, "linjer": count}
    except Exception as e:
        import traceback
        err = f"gem_mobilepay_dag fejl: {e}"
        traceback.print_exc()
        print(f"[ERROR] {err}")
        raise HTTPException(status_code=500, detail=err)


@app.get("/api/mobilepay/dag")
async def api_mobilepay_dag(request: Request, fra: str = None, til: str = None):
    _kræv_login(request)
    return database.hent_mobilepay_dag(fra, til)


@app.post("/api/mobilepay/upload-csv")
async def mobilepay_upload_csv(request: Request, fil: UploadFile = File(...)):
    """Browser CSV/Excel upload fra MobilePay portal."""
    _kræv_login(request)
    import tempfile, importlib
    from pathlib import Path as _Path
    suffix = _Path(fil.filename).suffix.lower() if fil.filename else ".csv"
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(await fil.read())
            tmp_path = tmp.name
        import mobilepay_csv_import as _mp_csv
        importlib.reload(_mp_csv)
        linjer = _mp_csv.parse_csv(tmp_path)
    except Exception as exc:
        print(f"[ERROR] CSV parse fejl: {exc}")
        import traceback
        traceback.print_exc()
        return {"ok": False, "fejl": str(exc)}
    finally:
        if tmp_path:
            try:
                import os as _os; _os.unlink(tmp_path)
            except OSError:
                pass
    if not linjer:
        return {"ok": True, "linjer": 0, "dage": 0, "total": 0, "total_gebyr": 0,
                "besked": "Ingen gyldige rækker fundet — tjek at filen er en Afregningsrapport eller Salgsoversigt fra portalen"}
    try:
        count = database.gem_mobilepay_dag(linjer)
        total_netto = sum(l.get("omsaetning_netto", l.get("omsaetning_inkl", 0)) for l in linjer)
        total_gebyr = sum(l.get("gebyr", 0) for l in linjer)
        return {
            "ok": True,
            "linjer": count,
            "dage": len(linjer),
            "total_netto": round(total_netto, 2),
            "total_gebyr": round(total_gebyr, 2),
            "total": round(total_netto + total_gebyr, 2)
        }
    except Exception as e:
        print(f"[ERROR] gem_mobilepay_dag fejl: {e}")
        import traceback
        traceback.print_exc()
        return {"ok": False, "fejl": f"Fejl ved gemning: {str(e)}"}


@app.get("/api/salg/mangler-kostpris")
async def api_mangler_kostpris(request: Request):
    _kræv_login(request)
    return database.hent_mangler_kostpris()


# ── TGTG ──────────────────────────────────────────────────────────────────────

@app.get("/api/tgtg/overblik")
async def api_tgtg_overblik(request: Request, aar: Optional[int] = None):
    _kræv_login(request)
    return database.hent_tgtg_overblik(aar)


@app.post("/api/tgtg/dagssalg")
async def api_tgtg_dagssalg(request: Request):
    """Modtager dagligt salg fra tgtg_sync.py eller manuel input."""
    body = await request.json()
    if body.get("secret") != WEBHOOK_SECRET:
        raise HTTPException(status_code=403, detail="Ugyldigt secret")
    linjer = body.get("linjer", [])
    if not linjer:
        raise HTTPException(status_code=400, detail="Ingen linjer")
    n = database.gem_tgtg_dagssalg(linjer)
    return {"ok": True, "linjer": n}


@app.post("/api/tgtg/poser")
async def api_tgtg_poser(request: Request):
    """Opdater pose-definitioner (navn + kreditpris)."""
    body = await request.json()
    if body.get("secret") != WEBHOOK_SECRET:
        raise HTTPException(status_code=403, detail="Ugyldigt secret")
    poser = body.get("poser", [])
    n = database.gem_tgtg_poser(poser)
    return {"ok": True, "poser": n}


@app.post("/api/tgtg/nulstil")
async def api_tgtg_nulstil(request: Request):
    """Slet al TGTG-data og pose-definitioner (til genindlæsning)."""
    body = await request.json()
    if body.get("secret") != WEBHOOK_SECRET:
        raise HTTPException(status_code=403, detail="Ugyldigt secret")
    with database._conn() as conn:
        conn.execute("DELETE FROM tgtg_dagssalg")
        conn.execute("DELETE FROM tgtg_poser")
    return {"ok": True, "besked": "Al TGTG-data slettet"}


# ── SPILD-RAPPORT ─────────────────────────────────────────────────────────────

@app.get("/api/spild/overblik")
async def api_spild_overblik(request: Request):
    """Spild-overblik for denne uge + forrige afsluttede uge til forside."""
    _kræv_login(request)
    from datetime import date
    iso  = date.today().isocalendar()
    uge, aar = iso[1], iso[0]
    prev_uge = uge - 1 if uge > 1 else 52
    prev_aar = aar if uge > 1 else aar - 1
    denne  = database.hent_spild_uge_overblik(uge,      aar)
    forrig = database.hent_spild_uge_overblik(prev_uge, prev_aar)
    return {"denne_uge": denne, "forrige_uge": forrig}


@app.get("/api/spild/dagsniveau")
async def api_spild_dagsniveau(request: Request, uge: Optional[int] = None, aar: Optional[int] = None):
    _kræv_login(request)
    if uge is None or aar is None:
        from datetime import date
        iso = date.today().isocalendar()
        uge = iso[1]
        aar = iso[0]
    return database.hent_spild_dagsniveau(int(uge), int(aar))


# ── VARESTAMDATA ──────────────────────────────────────────────────────────────

@app.get("/api/stamdata")
async def api_stamdata(request: Request):
    _kræv_login(request)
    return database.hent_stamdata()


@app.post("/api/stamdata/gem")
async def stamdata_gem(request: Request):
    _kræv_login(request)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Ugyldig JSON")
    varenavn = body.get("varenavn", "").strip()
    type_    = body.get("type", "").strip()
    if not varenavn or not type_:
        raise HTTPException(status_code=400, detail="Mangler varenavn eller type")
    id_ = database.gem_stamdata_linje(
        body.get("sku", ""),
        varenavn,
        type_,
        float(body.get("pris_ex_moms", 0) or 0),
        int(body.get("portioner", 1) or 1),
    )
    return {"ok": True, "id": id_}


@app.delete("/api/stamdata/{id_}")
async def stamdata_slet(request: Request, id_: int):
    _kræv_login(request)
    database.slet_stamdata(id_)
    return {"ok": True}


@app.post("/api/stamdata/bulk")
async def stamdata_bulk(request: Request):
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Ugyldig JSON")
    header_secret = request.headers.get("X-Webhook-Secret", "")
    if header_secret != WEBHOOK_SECRET and body.get("secret") != WEBHOOK_SECRET:
        raise HTTPException(status_code=401, detail="Ugyldig webhook secret")
    linjer = body.get("linjer", [])
    if not linjer:
        raise HTTPException(status_code=400, detail="Ingen linjer")
    antal = database.gem_stamdata_bulk(linjer)
    return {"ok": True, "linjer": antal}


@app.get("/api/kontrol/varenumre")
async def api_kontrol_varenumre(request: Request):
    _kræv_login(request)
    return database.hent_varenummer_kontrol()


@app.get("/api/debug/varer")
async def api_debug_varer(request: Request, q: str = "", bon: str = ""):
    """Debug: transaktioner + VF + stamdata for søgeord på seneste dato.
    Brug ?bon=NNNNN-... for at se alle varer på én specifik bon."""
    _kræv_login(request)
    with database._conn() as conn:
        seneste = conn.execute("SELECT MAX(dato) FROM transaktioner").fetchone()[0]
        if bon:
            # Vis alle rækker på en specifik bon
            bon_rækker = conn.execute("""
                SELECT bon_nr, varenavn, varenummer, antal, omsætning, kostpris
                FROM transaktioner
                WHERE bon_nr = ?
                ORDER BY varenavn
            """, (bon,)).fetchall()
            return {"bon": bon, "rækker": [dict(r) for r in bon_rækker]}
        raa = conn.execute("""
            SELECT varenavn, varenummer,
                   SUM(antal) AS antal,
                   SUM(omsætning) AS omsaetning,
                   SUM(kostpris) AS kostpris_shopbox
            FROM transaktioner
            WHERE dato = ? AND LOWER(varenavn) LIKE LOWER('%'||?||'%')
            GROUP BY varenavn, varenummer
        """, (seneste, q)).fetchall()
        view = conn.execute("""
            SELECT varenavn, varenummer,
                   SUM(antal) AS antal,
                   SUM(omsaetning_ex_moms) AS oms_ex,
                   SUM(vf_korrekt) AS vf,
                   SUM(db_korrekt) AS db
            FROM v_transaktioner
            WHERE dato = ? AND LOWER(varenavn) LIKE LOWER('%'||?||'%')
            GROUP BY varenavn, varenummer
        """, (seneste, q)).fetchall()
        stam = conn.execute("""
            SELECT sku, varenavn, pris_ex_moms, portioner
            FROM varestamdata
            WHERE LOWER(varenavn) LIKE LOWER('%'||?||'%')
               OR sku IN (SELECT DISTINCT varenummer FROM transaktioner
                          WHERE LOWER(varenavn) LIKE LOWER('%'||?||'%') AND varenummer != '')
        """, (q, q)).fetchall()
        rækker = conn.execute("""
            SELECT bon_nr, varenavn, varenummer, antal, omsætning, kostpris
            FROM transaktioner
            WHERE dato = ? AND LOWER(varenavn) LIKE LOWER('%'||?||'%')
            ORDER BY bon_nr, varenavn
        """, (seneste, q)).fetchall()
    return {
        "seneste_dato": seneste,
        "transaktioner": [dict(r) for r in raa],
        "v_transaktioner": [dict(r) for r in view],
        "stamdata": [dict(r) for r in stam],
        "rækker": [dict(r) for r in rækker],
    }


@app.get("/api/aarsplan/vf-detaljer")
async def api_vf_detaljer(request: Request, aar: int, maaned: int):
    _kræv_login(request)
    try:
        return database.hent_vf_detaljer(aar, maaned)
    except Exception as e:
        import traceback
        print(f"ERROR in vf_detaljer: {e}")
        traceback.print_exc()
        return {"error": str(e), "bager_vf": [], "andet_vf": []}


@app.get("/api/bager/fordelingsnoegle")
async def api_fordelingsnoegle(request: Request):
    """Vis den beregnede dagsnøgle til fordeling af bager-fakturaer."""
    _kræv_login(request)
    nøgle = database._dag_fordeling_nøgle()
    navne = ["Mandag","Tirsdag","Onsdag","Torsdag","Fredag","Lordag","Sondag"]
    return {
        "noegle": [
            {"dag": navne[i], "andel_pct": round(nøgle[i]*100, 1)}
            for i in range(7)
        ],
        "note": "Beregnet fra gennemsnitlig dagsomsaetning i transaktioner"
    }


@app.get("/api/debug/kager")
async def api_debug_kager(request: Request, uge: int, aar: int):
    """Debug: vis kage-varenavn fra bestillinger vs transaktioner for en uge."""
    _kræv_login(request)
    from datetime import date as _d, timedelta as _td
    jan4 = _d(aar, 1, 4)
    man  = jan4 - _td(days=jan4.weekday()) + _td(weeks=uge - 1)
    datoer = [(man + _td(days=i)).isoformat() for i in range(7)]
    ph = ','.join('?'*7)
    with database._conn() as conn:
        bestil = conn.execute("""
            SELECT varenavn, (man+tir+ons+tor+fre+loe+son) AS total
            FROM ugebestillinger WHERE uge=? AND aar=?
            AND (LOWER(varenavn) LIKE '%kage%' OR LOWER(varenavn) LIKE '%cookie%'
              OR LOWER(varenavn) LIKE '%muffin%' OR LOWER(varenavn) LIKE '%brownie%'
              OR LOWER(varenavn) LIKE '%romkugl%')
        """, (uge, aar)).fetchall()
        kasse = conn.execute(f"""
            SELECT varenavn, ROUND(SUM(antal),0) AS antal
            FROM transaktioner WHERE dato IN ({ph})
            AND (LOWER(varenavn) LIKE '%kage%' OR LOWER(varenavn) LIKE '%cookie%'
              OR LOWER(varenavn) LIKE '%muffin%' OR LOWER(varenavn) LIKE '%brownie%'
              OR LOWER(varenavn) LIKE '%romkugl%')
            GROUP BY varenavn
        """, datoer).fetchall()
    return {
        "bestillinger": [dict(r) for r in bestil],
        "kassesalg":    [dict(r) for r in kasse],
    }


@app.get("/api/faste-omk")
async def api_faste_omk(request: Request, aar: int):
    _kræv_login(request)
    return {"items": database.hent_faste_omk(aar)}


@app.post("/api/faste-omk/gem")
async def faste_omk_gem(request: Request):
    _kræv_login(request)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Ugyldig JSON")
    database.gem_faste_omk(
        int(body["aar"]),
        int(body["maaned"]),
        str(body["kategori"]),
        float(body["beloeb"]),
    )
    return {"ok": True}


@app.post("/api/faste-omk/slet-kategori")
async def faste_omk_slet(request: Request):
    _kræv_login(request)
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Ugyldig JSON")
    database.slet_faste_omk_kategori(int(body["aar"]), str(body["kategori"]))
    return {"ok": True}


@app.post("/api/bestilling/gem-manuel-wh")
async def bestilling_gem_manuel_webhook(request: Request):
    """Webhook-version: gem manuelle overrides uden login (kræver secret)."""
    body = await request.json()
    if body.get("secret") != WEBHOOK_SECRET:
        raise HTTPException(status_code=403, detail="Ugyldigt secret")
    uge  = body.get("uge")
    aar  = body.get("aar")
    overrides = body.get("overrides", [])  # [{varenummer, dag, antal}, ...]
    count = 0
    for o in overrides:
        vn   = o.get("varenummer")
        dag  = o.get("dag")
        antal = o.get("antal")
        if vn and dag and antal is not None:
            database.gem_bestilling_manuel(int(uge), int(aar), str(vn), str(dag), int(antal))
            count += 1
    return {"ok": True, "opdateret": count}


@app.get("/api/bestilling/anbefaling-wh")
async def api_bestilling_anbefaling_webhook(
    request: Request,
    secret: str = "",
    uge: Optional[int] = None,
    aar: Optional[int] = None,
):
    """Webhook-version af bestillingsanbefaling (ingen login krævet, kun secret)."""
    if secret != WEBHOOK_SECRET:
        raise HTTPException(status_code=403, detail="Ugyldigt secret")
    if uge is None:
        from datetime import date
        iso = date.today().isocalendar()
        uge = iso[1] + 1
        aar = iso[0]
        if uge > 52:
            uge = 1
            aar += 1
    if aar is None:
        from datetime import date
        aar = date.today().year
    return database.hent_bestillings_uge(int(uge), int(aar))


@app.post("/api/opdater-rapport")
async def opdater_rapport(request: Request):
    header_secret = request.headers.get("X-Webhook-Secret", "")

    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Ugyldig JSON")

    if header_secret != WEBHOOK_SECRET and body.get("secret") != WEBHOOK_SECRET:
        raise HTTPException(status_code=401, detail="Ugyldig webhook secret")

    encoded = body.get("data")
    if not encoded:
        raise HTTPException(status_code=400, detail="Ingen fil-data")

    dato_str = body.get("dato", datetime.now().isoformat())
    rapport_dato = dato_str[:10]

    try:
        fil_bytes     = base64.b64decode(encoded)
        transaktioner = xlsx_parser.parse_shopbox_xlsx(fil_bytes)

        if not transaktioner:
            raise HTTPException(status_code=422, detail="Ingen transaktioner fundet i filen")

        upload_id = database.gem_transaktioner(rapport_dato, transaktioner)
        return {"ok": True, "upload_id": upload_id, "rækker": len(transaktioner), "dato": rapport_dato}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Fejl ved behandling: {str(e)}")


# ── MANAGEMENT REVIEW ────────────────────────────────────────────────────────

@app.get("/api/management/review")
async def management_review_hent(request: Request):
    _kræv_login(request)
    data = database.hent_seneste_management_review()
    if not data:
        return {"ok": False, "ingen_data": True}
    return {"ok": True, "review": data}


@app.post("/api/management/review/opdater")
async def management_review_opdater(request: Request, uge: Optional[int] = None, aar: Optional[int] = None):
    _kræv_login(request)
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        raise HTTPException(status_code=503, detail="ANTHROPIC_API_KEY ikke sat i miljøvariable")
    try:
        review = database.generer_management_review(api_key, uge=uge, aar=aar)
        return {"ok": True, "review": review}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Fejl ved generering: {str(e)}")


@app.post("/api/management/spørg")
async def management_spørg(request: Request):
    _kræv_login(request)
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        raise HTTPException(status_code=503, detail="ANTHROPIC_API_KEY ikke sat")
    body = await request.json()
    spørgsmål = (body.get("spørgsmål") or "").strip()
    historik   = body.get("historik", [])
    if not spørgsmål:
        raise HTTPException(status_code=400, detail="Mangler spørgsmål")
    try:
        svar = database.besvar_data_spørgsmål(spørgsmål, historik, api_key)
        return {"ok": True, "svar": svar}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── VAREKOSTPRIS ──────────────────────────────────────────────────────────────

@app.get("/api/kostpris/oversigt")
async def api_kostpris_oversigt(request: Request):
    _kræv_login(request)
    return {"ok": True, "varer": database.hent_varekostpris_oversigt()}


@app.get("/api/kostpris/historik/{varenummer}")
async def api_kostpris_historik(request: Request, varenummer: str):
    _kræv_login(request)
    return {"ok": True, "historik": database.hent_varekostpris_historik(varenummer)}


@app.post("/api/kostpris/korriger")
async def api_kostpris_korriger(request: Request):
    _kræv_login(request)
    body = await request.json()
    vn   = str(body.get("varenummer", ""))
    pris = float(body.get("kostpris_enhed", 0))
    fra  = str(body.get("gyldig_fra", ""))
    if not vn or pris <= 0 or not fra:
        raise HTTPException(status_code=400, detail="varenummer, kostpris_enhed og gyldig_fra er påkrævet")
    database.korriger_varekostpris(vn, pris, fra)
    return {"ok": True}


# ── GMAIL AUTO-SYNC ───────────────────────────────────────────────────────────

@app.post("/api/bager/gmail-sync")
async def api_gmail_sync(request: Request):
    _kræv_login(request)
    result = gmail_sync_run()
    return result


@app.get("/api/bager/gmail-status")
async def api_gmail_status(request: Request):
    _kræv_login(request)
    status = database.hent_gmail_sync_status()
    har_token = bool(os.environ.get("GMAIL_TOKEN_JSON"))
    return {"ok": True, "status": status, "har_token": har_token}


# ── SALGSMØNSTER ─────────────────────────────────────────────────────────────

@app.get("/api/salg/sidst-solgt")
async def api_sidst_solgt(request: Request, uger: int = 4):
    _kræv_login(request)
    return database.hent_sidst_solgt_moenster(uger)


# ── VEJR ──────────────────────────────────────────────────────────────────────

@app.get("/api/vejr/forecast")
async def api_vejr_forecast(request: Request):
    _kræv_login(request)
    return database.hent_vejr_forecast()
