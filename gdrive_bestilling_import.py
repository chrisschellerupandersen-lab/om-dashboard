"""
Google Drive Ugebestilling Import
Henter automatisk ugebestillings-Excel fra Google Drive og importerer til Railway.

Kræver (én gang):
    pip install google-auth google-auth-oauthlib google-api-python-client openpyxl requests

Opsætning (første gang):
    python gdrive_bestilling_import.py --setup

Brug:
    python gdrive_bestilling_import.py              # seneste fil
    python gdrive_bestilling_import.py --uge 25     # specifik uge
    python gdrive_bestilling_import.py --vis        # vis uden upload
    python gdrive_bestilling_import.py --mappe ID   # specifik mappe-ID
"""

from __future__ import annotations
import argparse
import io
import json
import re
import sys
from datetime import date
from pathlib import Path
from typing import Optional

# ── Konfiguration ─────────────────────────────────────────────────────────────
RAILWAY_URL      = "https://om-dashboard-production-0f3a.up.railway.app"
WEBHOOK_SECRET   = "OM-Greve-2026-Hemlig"
# Leder efter credentials i denne rækkefølge:
# 1. gdrive_credentials.json (dedikeret)
# 2. gmail_credentials.json (samme Google Cloud projekt)
# 3. dashboard/ undermappen
_BASE = Path(__file__).parent
_DASH = _BASE / "dashboard"
CREDENTIALS_FILE = next(
    (p for p in [
        _BASE / "gdrive_credentials.json",
        _DASH / "gdrive_credentials.json",
        _BASE / "gmail_credentials.json",
        _DASH / "gmail_credentials.json",
    ] if p.exists()),
    _BASE / "gdrive_credentials.json"  # default (ikke-eksisterende → vejledning)
)
TOKEN_FILE = _BASE / "gdrive_token.json"

SCOPES = [
    "https://www.googleapis.com/auth/drive.readonly",
]

# Dag-kolonner i Excel (kolonner E-K = indeks 4-10)
DAGE_KEYS = ['man', 'tir', 'ons', 'tor', 'fre', 'loe', 'son']
# ─────────────────────────────────────────────────────────────────────────────


# ── Google Drive OAuth ────────────────────────────────────────────────────────

def _drive_service():
    """Returnerer autentificeret Google Drive API-service."""
    try:
        from google.oauth2.credentials import Credentials
        from google_auth_oauthlib.flow import InstalledAppFlow
        from google.auth.transport.requests import Request
        from googleapiclient.discovery import build
    except ImportError:
        print("[FEJL] Mangler pakker. Kør:")
        print("  pip install google-auth google-auth-oauthlib google-api-python-client")
        sys.exit(1)

    creds = None
    if TOKEN_FILE.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not CREDENTIALS_FILE.exists():
                print(f"[FEJL] Ingen credentials: {CREDENTIALS_FILE}")
                print()
                print("Opret Google Drive credentials:")
                print("  1. https://console.cloud.google.com/")
                print("  2. Aktiver 'Google Drive API'")
                print("  3. Opret OAuth 2.0 Client ID (Desktop)")
                print(f"  4. Download JSON og gem som: {CREDENTIALS_FILE}")
                print("  5. Kør igen")
                sys.exit(1)
            flow = InstalledAppFlow.from_client_secrets_file(str(CREDENTIALS_FILE), SCOPES)
            creds = flow.run_local_server(port=0)
        TOKEN_FILE.write_text(creds.to_json(), encoding="utf-8")

    return build("drive", "v3", credentials=creds)


# ── Find Excel-fil på Drive ───────────────────────────────────────────────────

def _find_bestilling_filer(service, mappe_id: Optional[str] = None, uge: Optional[int] = None) -> list[dict]:
    """Søg efter ugebestillings-Excel filer på Drive."""
    query_parts = [
        "mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'",
        "trashed=false",
    ]
    if mappe_id:
        query_parts.append(f"'{mappe_id}' in parents")

    # Søg på filnavne der matcher "bestilling" eller "uge"
    navn_query = "(name contains 'bestilling' or name contains 'Bestilling' or name contains 'uge' or name contains 'Uge')"
    query_parts.append(navn_query)

    query = " and ".join(query_parts)

    result = service.files().list(
        q=query,
        fields="files(id, name, modifiedTime, size)",
        orderBy="modifiedTime desc",
        pageSize=20
    ).execute()

    filer = result.get("files", [])

    # Filtrer på uge-nummer hvis angivet
    if uge:
        filer = [f for f in filer if str(uge) in f["name"]]

    return filer


def _download_fil(service, file_id: str) -> bytes:
    """Download fil-indhold."""
    from googleapiclient.http import MediaIoBaseDownload
    request = service.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    return buf.getvalue()


# ── Parse Excel ───────────────────────────────────────────────────────────────

def _parse_bestilling_xlsx(xlsx_bytes: bytes, filnavn: str) -> Optional[dict]:
    """
    Parser ugebestillings-Excel og returnerer {uge, aar, linjer}.
    Forventer samme format som bestilling_template.xlsx:
      - A2: "Uge XX"
      - Kolonne B: varenummer, C: kategori, D: pris, E-K: man-søn
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[FEJL] openpyxl ikke installeret: pip install openpyxl")
        sys.exit(1)

    # Find uge fra filnavn
    uge = None
    aar = date.today().year
    m = re.search(r'uge\s*(\d+)', filnavn, re.IGNORECASE)
    if m:
        uge = int(m.group(1))
    m2 = re.search(r'(20\d\d)', filnavn)
    if m2:
        aar = int(m2.group(1))

    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        ws = wb.active

        # Prøv at finde uge fra celle A2
        a2 = ws['A2'].value
        if a2 and isinstance(a2, str):
            m3 = re.search(r'uge\s*(\d+)', a2, re.IGNORECASE)
            if m3:
                uge = int(m3.group(1))

        if not uge:
            print(f"[ADVARSEL] Kunne ikke finde ugenummer i '{filnavn}' — brug --uge parameter")
            return None

        linjer = []
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
            # Forvent: [rækkenr, varenummer, varenavn, kategori, pris, man, tir, ons, tor, fre, loe, son, total, ...]
            # Tjek at rækken har data
            if not row or not row[1]:
                continue

            varenummer = row[1]
            if not isinstance(varenummer, (int, float)):
                continue

            varenavn = row[2] or ""
            kategori = row[3] if len(row) > 3 else ""
            pris     = float(row[4] or 0) if len(row) > 4 else 0.0

            # Dag-værdier E-K (indeks 5-11)
            dag_vals = {}
            for i, dk in enumerate(DAGE_KEYS):
                col_idx = 5 + i
                val = row[col_idx] if len(row) > col_idx else None
                dag_vals[dk] = int(float(val)) if val and val != '' else 0

            total = sum(dag_vals.values())
            if total == 0:
                continue  # Spring tomme rækker over

            linjer.append({
                "varenummer":   str(int(varenummer)),
                "varenavn":     str(varenavn).strip(),
                "kategori":     str(kategori).strip() if kategori else "",
                "pris_ex_moms": pris,
                **dag_vals,
                "total_antal":  total,
                "total_pris":   round(total * pris, 2),
            })

        return {"uge": uge, "aar": aar, "linjer": linjer, "filnavn": filnavn}

    except Exception as e:
        print(f"[FEJL] Parse-fejl: {e}")
        return None


# ── Upload til Railway ─────────────────────────────────────────────────────────

def _upload(data: dict) -> bool:
    """POST bestillingslinjer til Railway."""
    try:
        import requests
    except ImportError:
        print("[FEJL] requests ikke installeret: pip install requests")
        sys.exit(1)

    payload = {
        "secret": WEBHOOK_SECRET,
        "uge":    data["uge"],
        "aar":    data["aar"],
        "linjer": data["linjer"],
    }
    r = requests.post(
        f"{RAILWAY_URL}/api/bestilling/opdater",
        json=payload,
        timeout=30,
    )
    r.raise_for_status()
    res = r.json()
    print(f"[OK] Uge {data['uge']}/{data['aar']} — {res.get('linjer','?')} linjer uploadet")
    return True


# ── Vis resultat ──────────────────────────────────────────────────────────────

def _vis(data: dict) -> None:
    print(f"\n  Fil:  {data['filnavn']}")
    print(f"  Uge:  {data['uge']}/{data['aar']}")
    print(f"  Linjer: {len(data['linjer'])}")
    print()
    for l in data["linjer"][:10]:
        dage_str = ' '.join([f"{dk}:{l[dk]}" for dk in DAGE_KEYS if l.get(dk, 0) > 0])
        print(f"  {l['varenummer']:>6}  {l['varenavn'][:30]:<30}  {dage_str}")
    if len(data["linjer"]) > 10:
        print(f"  ... og {len(data['linjer'])-10} mere")


# ── Hoved-flow ────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Google Drive Ugebestilling → Railway")
    parser.add_argument("--setup",  action="store_true", help="Opret Google OAuth token")
    parser.add_argument("--uge",    type=int,            help="Specifik uge")
    parser.add_argument("--mappe",  default=None,        help="Google Drive mappe-ID")
    parser.add_argument("--vis",    action="store_true", help="Vis uden upload")
    parser.add_argument("--fil",    default=None,        help="Specifik fil-ID på Drive")
    args = parser.parse_args()

    service = _drive_service()

    if args.setup:
        print("[OK] Google Drive OAuth oprettet.")
        return

    # Find fil
    if args.fil:
        # Direkte fil-ID
        fil_meta = service.files().get(fileId=args.fil, fields="id,name,modifiedTime").execute()
        filer = [fil_meta]
    else:
        filer = _find_bestilling_filer(service, mappe_id=args.mappe, uge=args.uge)

    if not filer:
        print("[INFO] Ingen bestillings-Excel fundet på Drive.")
        print("  Tips:")
        print("  - Brug --mappe MAPPE_ID for at angive specifik mappe")
        print("  - Brug --fil FIL_ID for specifik fil")
        print("  - Filnavnet skal indeholde 'bestilling' eller 'uge'")
        return

    print(f"[INFO] Fandt {len(filer)} fil(er):")
    for i, f in enumerate(filer):
        print(f"  [{i+1}] {f['name']} (ændret: {f['modifiedTime'][:10]})")

    # Brug seneste fil (eller den der matcher uge)
    valgt = filer[0]
    print(f"\n[INFO] Henter: {valgt['name']}")

    xlsx_bytes = _download_fil(service, valgt["id"])
    data = _parse_bestilling_xlsx(xlsx_bytes, valgt["name"])

    if not data:
        print("[FEJL] Kunne ikke parse filen.")
        return

    if not data["linjer"]:
        print("[ADVARSEL] Ingen bestillingslinjer fundet i filen.")
        return

    _vis(data)

    if args.vis:
        print("\n[VIS] Ingen upload (--vis flag)")
        return

    svar = input(f"\nUpload uge {data['uge']}/{data['aar']} ({len(data['linjer'])} linjer) til Railway? [J/n] ").strip().lower()
    if svar not in ("", "j", "ja", "y", "yes"):
        print("Afbrudt.")
        return

    _upload(data)


if __name__ == "__main__":
    main()
